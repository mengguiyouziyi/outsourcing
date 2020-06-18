import pathlib
from openpyxl import load_workbook
from openpyxl.drawing.image import Image, _import_image
import os
import time

##1106100(4)_PHOTOINSERT ,INPUT AND ROW AND 间隔行都可以选择20200602

if not os.path.exists('./picpath.txt') and not os.path.exists(r'c:\photo'):
    print("[-] 没有图片路径")
    exit()


def find_pic(id):
    switch_dir = 0
    try:
        with open('picpath.txt') as f:
            pic_dir = pathlib.Path(f.readline().strip())
    except FileNotFoundError:
        pic_dir = pathlib.Path(r"C:\photo")
        switch_dir = 1
    while True:
        result_jpg = list(pic_dir.glob(f'{product_id}.jpg'))
        result_png = list(pic_dir.glob(f'{product_id}.png'))
        if result_jpg and result_png:
            print(f"\t\t[-] 找到同名不同后缀的图片！ {result}")
            pic_path = None
        elif len(result_jpg) > 0:
            pic_path = result_jpg[0]
        elif len(result_png) > 0:
            pic_path = result_png[0]
        else:
            print(f"\t\t[-] 找不到图片{product_id}.jpg 或者 {product_id}.png！")
            pic_path = None
        if pic_path:
            return pic_path
        elif switch_dir == 0:
            pic_dir = pathlib.Path(r"C:\photo")
            switch_dir = 1
        else:
            return None


keyword = input("请输入关键字：")
column_name = input("请输入列：")
line_num = int(input("请输入间隔行数："))

p = pathlib.Path('.')

files = list(p.glob("*.xlsx"))
print(files)
count = 0
is_ok = False
for f in files:
    wb = load_workbook(str(f))
    print(wb)
    ws = wb['Sheet1']
    if ws['A1'].value == '图片处理':
        count += 1
        print(f"[*] 正在处理表格{f}")
        for ws in wb:
            if ws.title == 'Sheet1':
                pass
            else:
                print(f"\t[*] 正在处理表单{ws.title}")
                # 寻找产品编号所在单元格
                is_found = False
                for r in ws.rows:
                    for cell in r:
                        if type(cell.value) == str and cell.value.upper() == keyword.upper():
                            row, column = cell.row + 1, cell.column_letter
                            is_found = True
                            is_ok = True
                            break
                if not is_found:
                    print("\t[*] 没有找到关键字")
                    continue
                while row <= ws.max_row:
                    product_id = ws[f'{column}{row}'].value
                    pic_path = find_pic(product_id)
                    if pic_path:
                        print("\t\t[*]", pic_path)
                        pic = Image(str(pic_path))
                        height = ws.row_dimensions[row].height
                        if height == None:
                            height = 15
                        pic.height, pic.width = height, pic.width * height / pic.height
                        ws.add_image(pic, f'{column_name}{row}')
                    else:
                        print("\t\t[*] 没有找到！")
                    row += (line_num + 1)
        if is_ok:
            wb.save(str(f))
            print(f"[*] 保存到{f}")
        else:
            print("没有符合要求的Sheet！")
            time.sleep(3)

    else:
        continue
