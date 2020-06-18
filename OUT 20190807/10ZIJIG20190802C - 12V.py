from openpyxl import load_workbook
import time
import os


def cmp(file):
    return time.strptime(file.split('_')[1], '%Y%m%d%H%M')


files = [i for i in os.listdir('.') if i.startswith('PAYMENT')]

files = sorted(files, key=cmp)

excel_name = files[-1]

print(f"正在处理 {excel_name}")


def get_current_time(f='%Y%m%d%H%M'):
    t = time.time()
    localtime = time.localtime(t)
    return time.strftime(f, localtime)


def handle(sheet, is_sheet3=False):
    wb = load_workbook(filename=excel_name)
    wb_data = load_workbook(filename=excel_name, data_only=True)
    inv_sheets = [i for i in wb.worksheets if i.title.lstrip().startswith('INV')]
    sheet2 = wb[sheet]

    sheet2_row = 4
    total = 0
    count = 0
    for inv in inv_sheets:
        count += 1
        sheet2[f'A{sheet2_row}'].value = f'#{inv.title}'
        sheet2_row = sheet2_row + 1
        inv_row = 3
        good_values = 0
        payment = 0
        while inv[f'B{inv_row}'].value != None:
            if inv[f'F{inv_row}'].value == '*' and is_sheet3:
                inv_row += 1
                continue
            for i in 'ABCD':
                sheet2[f'{i}{sheet2_row}'].value = inv[f'{i}{inv_row}'].value
                if i == 'B' and is_sheet3:
                    pos1 = sheet2[f'{i}{sheet2_row}'].value.find('*')
                    pos2 = sheet2[f'{i}{sheet2_row}'].value.find('＊')
                    if pos1 > 0:
                        sheet2[f'{i}{sheet2_row}'].value = sheet2[f'{i}{sheet2_row}'].value[:pos1]
                    elif pos2 > 0:
                        sheet2[f'{i}{sheet2_row}'].value = sheet2[f'{i}{sheet2_row}'].value[:pos2]
            if inv[f'C{inv_row}'].value:
                good_values += inv[f'C{inv_row}'].value
            if inv[f'D{inv_row}'].value:
                payment += inv[f'D{inv_row}'].value
            inv_row += 1
            sheet2_row += 1
        sheet2[f'A{sheet2_row}'].value = '#'
        sheet2[f'F{sheet2_row}'].value = payment - good_values
        total += payment - good_values
        sheet2_row += 1
    sheet2[f'G{sheet2_row-1}'].value = total
    sheet1 = wb_data['Sheet1']
    sheet2['A1'].value = sheet1['B41'].value
    ##    sheet2['B1'].value = sheet1['D41'].value + sheet1['F41'].value
    sheet2['C1'].value = sheet1['H41'].value
    sheet2['D1'].value = get_current_time()
    sheet2['A3'].value = f'{count} INV' if is_sheet3 else f'{count}INV'
    sheet1 = wb['Sheet1']
    if is_sheet3:
        name = f"ACCOUNT_{sheet1['B2'].value}_{get_current_time()}_{count}INV_{total}.xlsx"
    else:
        name = f"ACCOUNT_{sheet1['B4'].value}_{sheet1['D4'].value}_{sheet1['F4'].value}_{get_current_time()}_{count}单_{total}.xlsx"
    tail = 1
    new_name = name
    while os.path.exists(new_name):
        new_name = name.split('.')[0] + f'-{tail}.xlsx'
        tail += 1
    sheet_new_name = f"TOTAL INV {sheet1['B2'].value} {sheet1['B7'].value} {get_current_time('%Y%m%d')}"
    for s in wb.sheetnames:
        if s != sheet:
            wb.remove(wb[s])
    sheet2.title = sheet_new_name
    wb.save(new_name)
    print(f"新的文件为：{new_name}")


handle('sheet 2')
handle('sheet 3', True)
